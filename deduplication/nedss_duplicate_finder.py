"""NEDSS data duplicate person-record detector.

This program automates the process of identifying potential duplicate person-records
in NEDSS data. The program takes two command line arguments (1) the filepath
for NEDSS data; (2) the Identifier of the first person-record considered new.
This application will then compare each of the new records to all the preceeding records.

Example: for a new batch of NEDSS data, we need to determine the Identifier value for
the first record in that dataset that should be considered new. E.g., we may open the
NEDSS data file and determine that the new data starts at Identifier value 11960. Thus:

    py nedss_duplicate_person_record_detector.py "09-29-20_Positive Cases.xlsx" -i 11960

The output from the above command is saved as "09-29-20_Positive Cases_w_DUPE_UUID.xlsx"
In this file the column dupe_uuid stores potential duplicate identifiers. Filtering
this data to only the records that contains non-empty values (and sorting on dupe_uuid)
will present an ordered list of potential dupes for manual review.

"""

import argparse  # for command line arg parsing

from collections import Counter
import jellyfish  # for phonetic representation
import openpyxl as xl  # for writing to xls file
import pandas as pd  # for data management/aggs.
from scipy import sparse  # for connected comps.
import textdistance  # for fuzzy string matching
from tqdm import tqdm  # for process status bar
from typing import List


def first_name_similarity_scorer(a: str, b: str):
    """Compares two first name strings, returns 1 if they match, 0 otherwise.
    
    Uses Jaro-Winkler Distance Algorithm (JWDA) (en.wikipedia.org/wiki/Jaro–Winkler_distance).
    JWDA "measurement scale is 0.0 to 1.0, where 0.0 is the least likely and 1.0 is a positive match.  
    For our purposes, anything below a 0.8 is not considered useful." (source: SAP blog)
    
    """
    jaro = textdistance.jaro_winkler(a, b)
    if jaro > 0.8:
        return 1
    else:
        return 0


def last_name_similarity_scorer(a: str, b: str):
    """Compares two first last strings, returns 1 if they match, 0 otherwise.
    
    Uses Jaro-Winkler Distance Algorithm (JWDA) (en.wikipedia.org/wiki/Jaro–Winkler_distance).
    JWDA "measurement scale is 0.0 to 1.0, where 0.0 is the least likely and 1.0 is a positive match.  
    For our purposes, anything below a 0.8 is not considered useful." (source: SAP blog)
    
    """
    jaro = textdistance.jaro_winkler(a, b)
    if jaro > 0.8:
        return 1
    else:
        return 0


def address_similarity_scorer(a: str, b: str):
    """Compares two address strings, returns 1 if they match, 0 otherwise.
    
    Uses Jaro-Winkler Distance Algorithm (JWDA) (en.wikipedia.org/wiki/Jaro–Winkler_distance).
    JWDA "measurement scale is 0.0 to 1.0, where 0.0 is the least likely and 1.0 is a positive match.  
    
    """
    jaro = textdistance.jaro_winkler(a, b)
    if jaro > 0.9:
        return 1
    else:
        return 0


def age_similarity_scorer(age_1, age_2):
    """Compares two ages, returns 0 if they match, returns penalty of -1 otherwise.

    Conservative assumptions:
        1) If age cannot be cleanly cast to int, consider comparators to be a match
        2) If ages are within 2 years, consider comparators to be a match
    
    """
    try:
        age_1 = int(age_1)
    except:
        return 0
    try:
        age_2 = int(age_2)
    except:
        return 0

    # client requested age tolerance of 2 years:
    if abs(age_1 - age_2) <= 2:
        return 0
    else:
        return -1


def gender_similarity_scorer(gender_1: str, gender_2: str):
    """Compares two gender strings, returns 0 if they match, returns penalty of -1 otherwise.

    Conservative assumption: if gender is nil or empty string, consider it a match against the comparator.
    """
    if gender_1 == gender_2:
        return 0
    elif gender_1 is None or gender_2 is None:
        return 0
    elif gender_1 == "" or gender_2 == "":
        return 0
    else:
        return -1


def preprocess(nedss_df: pd.DataFrame):
    """Translates name and address fields into phonetic representations."""
    nedss_df.columns = nedss_df.columns.str.strip()
    nedss_df["phonetic_first_name"] = nedss_df["First Name"].apply(jellyfish.metaphone)
    nedss_df["phonetic_last_name"] = nedss_df["Last Name"].apply(jellyfish.metaphone)
    nedss_df["phonetic_address"] = nedss_df["Address"].apply(jellyfish.metaphone)
    return nedss_df


def score_record_against_records(nedss_df: pd.DataFrame, nedss_row: pd.Series) -> pd.DataFrame:
    """Calculates various similarity scores and total score, then returns the score dataframe."""
    score_df = pd.DataFrame(index=nedss_df.index)
    score_df['first_name_similarity_score'] = nedss_df['phonetic_first_name'].apply(
            first_name_similarity_scorer, args=(nedss_row['phonetic_first_name'],))
    score_df['last_name_similarity_score'] = nedss_df['phonetic_last_name'].apply(
            last_name_similarity_scorer, args=(nedss_row["phonetic_last_name"],))
    score_df['address_similarity_score'] = nedss_df['phonetic_address'].apply(
            address_similarity_scorer, args=(nedss_row["phonetic_address"],))
    score_df['age_similarity_score'] = nedss_df["Age"].apply(
            age_similarity_scorer, args=(nedss_row["Age"],))
    score_df['gender_similarity_score'] = nedss_df["Gender"].apply(
            gender_similarity_scorer, args=(nedss_row["Gender"],))
    score_df["total_score"] = score_df.sum(axis=1)
    return score_df


def get_dupe_groups(adj_mat) -> List[List[int]]:
    """Takes an adjacency matrix, applies connected components to link potential dupe records."""
    dupe_lists = []
    n_components, labels = sparse.csgraph.connected_components(
            csgraph=adj_mat, directed=False, return_labels=True)
    # get the connected components (with more than one member):
    dupe_ids = [k for k, v in Counter(labels).items() if v > 1]
    # for each connected component (with more than one member) find the indices for its members:
    for dupe_id in dupe_ids:
        dupe_lists.append([i for i, x in enumerate(labels) if x == dupe_id])
    return dupe_lists


def get_dupe_index_groups(data_df: pd.DataFrame, split_index: int) -> List[List[int]]:
    """Takes dataframe to process and index first new record, and returns list of linked records"""
    mat_dim = data_df.index.max() + 1
    adj_mat = sparse.dok_matrix((mat_dim, mat_dim), dtype='int64')
    new_record_indices = [i for i in data_df.index.tolist() if i >= split_index]
    for rec_idx in tqdm(new_record_indices):
        df_old = data_df.loc[:rec_idx - 1]
        df_new_row = data_df.loc[rec_idx]
        score_df = score_record_against_records(df_old, df_new_row)
        match_df = score_df[score_df["total_score"] > 1]
        for match_idx in match_df.index:
            adj_mat[rec_idx, match_idx] = 1
    return get_dupe_groups(adj_mat)


def write_dupe_info_to_workbook(in_filepath: str, out_filepath: str, dupe_lists: List[List[int]]):
    """Takes a list of lists of duplicate record indices, annotates input file with dupe ids."""
    wb = xl.load_workbook(in_filepath)
    ws = wb.worksheets[0]
    new_col_idx = ws.max_column + 1
    ws.cell(row=1, column=new_col_idx, value="dupe_group")
    for i, dupe_group in enumerate(dupe_lists):
        for dupe in dupe_group:
            # +2 for header and 1-index offset
            ws.cell(row=dupe + 2, column=new_col_idx, value=i)
    if not out_filepath:
        out_filepath = in_filepath.split(".")
        out_filepath[-2] = out_filepath[-2] + "_w_DUPE_UUID"
        out_filepath = ".".join(out_filepath)
    wb.save(out_filepath)


if __name__ == "__main__":
    arg_parser = argparse.ArgumentParser(description='Flag NEDSS data for review')
    arg_parser.add_argument('-p',
                            '--path',
                            # metavar='path',
                            type=str,
                            help='the path to the raw NEDSS data file')
    arg_parser.add_argument('-i',
                            '--identifier',
                            type=int,
                            # required=True,
                            help='Identifier for first new record in NEDSS data')
    arg_parser.add_argument('-o',
                            '--output',
                            type=str,
                            help='where to save the output file')
    args = arg_parser.parse_args()

    path = args.path
    output = args.output
    identifier = args.identifier

    if not path or not output or not identifier:
        path = input("Paste the input file path here (then press enter): ")
        path = path.strip('"')
        output = input("Paste the output file path here (then press enter): ")
        output = output.strip('"')
        identifier = input("Type the Identifier for the first new record here (then press enter): ")
        while identifier == "":
            identifier = input("Type the Identifier for the first new record here (then press enter): ")
        if identifier == "1" or identifier == "0":
            identifier = "2"
        identifier = int(identifier)

    df = pd.read_excel(path)
    df = preprocess(df)  # do all up-front preprocessing here (e.g., phonetics)
    index_of_first_new_record = df[df["Identifier"] == identifier].index[0]
    dupe_groups = get_dupe_index_groups(df, index_of_first_new_record)

    write_dupe_info_to_workbook(path, output, dupe_groups)
